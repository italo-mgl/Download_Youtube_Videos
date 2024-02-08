from pytube import YouTube
import os
from openpyxl import load_workbook

# Função para baixar o áudio em formato MP3 de um vídeo do YouTube
def baixar_audio_mp3(url):
    try:
        yt = YouTube(url)
        stream = yt.streams.filter(only_audio=True).first()
        output_path = os.path.join("downloads", f"{yt.title}.mp3")
        stream.download(output_path=output_path)
        print(f"Vídeo baixado como áudio MP3: {yt.title}.mp3")
    except Exception as e:
        print(f"Erro ao baixar o vídeo {url}: {e}")

# Carregar a planilha
workbook = load_workbook(filename="trunks_Futuro.xlsx")
sheet = workbook.active

# Iterar sobre as células da planilha e obter os links dos vídeos
for row in sheet.iter_rows(values_only=True):
    for cell in row:
        url = str(cell).strip()
        if "youtube.com" in url:
            print(f"Baixando vídeo do link: {url}")
            baixar_audio_mp3(url)
                        baixar_audio_mp3(url)

